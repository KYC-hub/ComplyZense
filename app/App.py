# to host the app, run the command after starting app: 
# ngrok http http://localhost:5000
# ngrok http --url=engaged-sunfish-terribly.ngrok-free.app 5000
import io
from io import BytesIO
import re
import secrets
from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    session,
    send_file,
)
from flask_cors import CORS
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pytesseract
from PIL import Image
import pdfplumber
from docx import Document
import pandas as pd
import openpyxl
import os
import openai
import json
import csv
import html
import uuid
from datetime import datetime, timedelta
from dotenv import load_dotenv

# Load environment details from .env file
load_dotenv()
API_KEY = os.getenv("OPENAI_API_KEY")
MODEL_ID = os.getenv("OPENAI_MODEL_ID")
os.environ["OPENAI_API_KEY"] = API_KEY

# Initialize Flask app
app = Flask(__name__)
CORS(app)
UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.secret_key = secrets.token_hex(16)  # To ensure session security
app.config["ALLOWED_EXTENSIONS"] = {
    "txt",
    "csv",
    "docx",
    "jpg",
    "jpeg",
    "png",
    "pdf",
    "xlsx",
    "reg",
    "md",
    "json",
}
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024

# Database connection function
def get_db_connection():
    try:
        conn = sqlite3.connect("database.db")
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON") 
        return conn
    except sqlite3.Error as e:
        print(f"Database connection error: {e}")
        return None
    
# Initialize the database
def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Create table for users if it does not exist
    cursor.execute(
        """CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )"""
    )

    # Create table for chat_sessions (with ON DELETE CASCADE)
    cursor.execute(
        """CREATE TABLE IF NOT EXISTS chat_sessions (
            session_id TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            session_name TEXT DEFAULT '1',
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )"""
    )

    # Create table for chat messages (with ON DELETE CASCADE)
    cursor.execute(
        """CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            message TEXT,
            response TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES chat_sessions(session_id) ON DELETE CASCADE
        )"""
    )

    conn.commit()
    conn.close()
    print("Database initialized successfully!")
    
# Function to get UTC+8 time
def get_utc_plus_8_time():
    utc_now = datetime.utcnow()
    utc_plus_8 = utc_now + timedelta(hours=8)
    # Format with 12-hour clock and AM/PM
    return utc_plus_8.strftime('%Y-%m-%d %I:%M:%S %p')

# Function to add a new chat session
def add_chat_session(user_id, session_name):
    conn = get_db_connection()
    cursor = conn.cursor()

    timestamp = get_utc_plus_8_time()
    
    cursor.execute(
        """INSERT INTO chat_sessions (session_id, user_id, session_name, timestamp) 
        VALUES (?, ?, ?, ?)""",
        (str(uuid.uuid4()), user_id, session_name, timestamp)
    )
    conn.commit()
    conn.close()

# Route to check if user is logged in
@app.route("/check_login", methods=["GET"])
def check_login():
    if "user_id" in session:
        user_id = session["user_id"]
        username = session["username"]

        # Fetch session_name from the session
        session_name = session.get("session_name", "No active session")

        return jsonify({
            "isLoggedIn": True,
            "username": username,
            "sessionname": session_name 
        })
    
    return jsonify({
        "isLoggedIn": False,
        "username": None,
        "sessionname": None 
    })

# Route for user login
@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        # Basic validation
        if not username or not password:
            return render_template("login.html", error="Please fill in all fields")

        username = html.escape(username)

        # Input length validation
        if len(username) < 3 or len(username) > 50:
            return render_template(
                "login.html", error="Username must be between 3 and 50 characters"
            )

        try:
            # Get user data
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()
            conn.close()

            # Authenticate user
            if user and check_password_hash(user["password"], password):
                # Clear existing session
                session.clear()

                # Store user data in session
                session.update({
                    "user_id": user["id"],
                    "username": user["username"],
                })

                # Get the highest session_name
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT MAX(session_name) FROM chat_sessions WHERE user_id = ?",
                    (user["id"],)
                )
                max_session_name = cursor.fetchone()[0]
                conn.close()

                # Determine next session_name
                session_name = 1 if max_session_name is None else max_session_name + 1
                session_id = str(uuid.uuid4())
                timestamp = get_utc_plus_8_time()

                # Store session in database
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO chat_sessions (session_id, user_id, session_name, timestamp) VALUES (?, ?, ?, ?)",
                    (session_id, user["id"], session_name, timestamp),
                )
                conn.commit()
                conn.close()

                # Update session with session details
                session["session_id"] = session_id
                session["session_name"] = session_name
                
                return redirect(url_for("chatbot"))
            else:
                return render_template("login.html", error="Invalid credentials")

        except Exception as e:
            print(f"Error during login: {e}")
            return render_template(
                "login.html", error="An error occurred. Please try again later."
            )
    return render_template("login.html")

# Route for registration page
@app.route("/register", methods=["GET", "POST"])
def register_page():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        name = request.form.get("name")

        if not username or not password or not name:
            return render_template("register.html", error="Please fill in all fields")

        # Escape special characters to prevent XSS
        username = html.escape(username)
        name = html.escape(name)

        # Input length validation
        if len(username) < 3 or len(username) > 50:
            return render_template(
                "register.html", error="Username must be between 3 and 50 characters"
            )

        if len(password) < 6:
            return render_template(
                "register.html", error="Password must be at least 6 characters long"
            )

        if len(name) < 2 or len(name) > 100:
            return render_template(
                "register.html", error="Name must be between 2 and 100 characters"
            )

        # Check username format for alphanumeric characters or other requirements
        if not re.match("^[a-zA-Z0-9_]*$", username):
            return render_template(
                "register.html", error="Username must be alphanumeric"
            )

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Check if the username already exists
            cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
            existing_user = cursor.fetchone()
            if existing_user:
                return render_template("register.html", error="Username already exists")

            # Hash the password and insert the new user
            hashed_password = generate_password_hash(password)
            cursor.execute(
                "INSERT INTO users (name, username, password) VALUES (?, ?, ?)",
                (name, username, hashed_password),
            )
            conn.commit()
            conn.close()

            # After registration, redirect to login page
            return redirect(url_for("login_page"))

        except Exception as e:
            print(f"Error during registration: {e}")
            return render_template(
                "register.html", error="An error occurred. Please try again later."
            )

    return render_template("register.html")

# Main route (chatbot page)
@app.route("/", methods=["GET"])
def chatbot():
    return render_template("chatbot.html")

def get_conversation_history(session_id, limit=20):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT message, response FROM messages WHERE session_id = ? ORDER BY timestamp DESC LIMIT ?",
            (session_id, limit),
        )
        rows = cursor.fetchall()
        conn.close()
        return rows[::-1]  # Reverse the order to show the latest messages first
    except Exception as e:
        print(f"Error getting conversation history: {e}")
        return []

prompt_instructions = """
You are an AI assistant specialized in providing detailed and accurate information on ISO 27001 and ISO 27002 standards, specifically the 2022 editions.
Your goal is to help users understand the principles, requirements, and best practices for implementing and managing information security management systems (ISMS).
You should provide clear, concise, and authoritative responses based on the ISO 27001 and ISO 27002 frameworks, while considering their latest updates and revisions.
Ensure your answers are aligned with the principles of information security, such as confidentiality, integrity, and availability, and be mindful to clarify any ambiguous terms or concepts related to these standards.
You can also offer practical guidance on risk management, controls, and audits based on the ISO 27001 and 27002 guidelines.
Be objective, neutral, and professional, and avoid providing personal opinions or legal advice.
"""

# Function to get GPT response
def get_response(message):
    current_session_id = session.get("session_id")  
    
    try:
        history = get_conversation_history(current_session_id)
        
        if history:
            chat_history = "\n".join([f"user: {row[0]}\nassistant: {row[1]}" for row in history])
            message = f"{chat_history}\nuser: {message}"
        
        chat_completion = openai.chat.completions.create(
            model=MODEL_ID,
            messages=[
                {
                    "role": "system",
                    "content": prompt_instructions,
                },
                {"role": "user", "content": message},
            ],
        )
        reply = chat_completion.choices[0].message.content
        return reply
    except openai.OpenAIError as e:
        print("OpenAI API error:", e)

report_instructions = """
You are a compliance and audit reporting assistant. 
Your task is to analyze the provided company policies and generate a structured, professional compliance report aligned with ISO 27001 and ISO 27002 (2022 editions).
Your report should identify compliance issues, assess policy quality, and provide actionable recommendations.

### Report Guidelines:

#### Title:
Begin with a clear, relevant title (e.g., "Compliance and Policy Quality Review").

#### Structure:
For each policy or policy area, include:
- Strengths Identified: Highlight effective measures and compliance with best practices.
- Issues Detected: Identify gaps, risks, or non-compliance concerns.
- Impact: Explain potential vulnerabilities and risks associated with each issue.
- Recommendations: Provide specific, actionable solutions to address compliance gaps.

#### Policy Quality Evaluation:
- Good Policy Attributes: Clearly defined terms, strong access control, regular reviews, alignment with standards.
- Bad Policy Attributes: Vague language, lack of accountability, outdated procedures.
- Overall Assessment: Classify the policy as "Good" or "Needs Improvement" (or use a rating system) with a brief justification.

#### Compliance Issues Format:
For each compliance issue, use this format:
- Description: Explain the issue and its root cause.
- Severity Level: High / Medium / Low.
- Impact: Describe the risks or vulnerabilities.
- Solutions: Provide 1–3 actionable recommendations.

### Tone and Style:
- Maintain a formal, precise, and professional tone.
- Avoid unnecessary introductory phrases (e.g., "It appears that...").
- Reference specific sections using **identifiers** (e.g., section numbers or headings) for clarity.

### Objective:
Deliver a concise, actionable compliance report that helps decision-makers understand strengths, identify weaknesses, and implement improvements for stronger compliance and policy effectiveness.
"""

# Function to generate report
def get_report(message):
    try:
        chat_completion = openai.chat.completions.create(
            model=MODEL_ID,
            temperature = 0.1, # Ensures no randomness for report generation
            messages=[
                {
                    "role": "system",
                    "content": report_instructions,
                },
                {"role": "user", "content": message},
            ],
        )
        reply = chat_completion.choices[0].message.content
        return reply
    except openai.OpenAIError as e:
        print("OpenAI API error:", e)

# Function to check if the file extension is allowed
def allowed_file(filename):
    allowed_extensions = app.config.get("ALLOWED_EXTENSIONS", set())
    if not filename or "." not in filename:
        return False
    # Extract file extension
    extension = filename.rsplit(".", 1)[1].lower()
    return extension in allowed_extensions

# Extract text from PDF
def extract_text_from_pdf(file_path):
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    lines = page_text.split("\n")
                    temp_text = ""
                    for line in lines:
                        line = line.strip()
                        if re.search(r"[.!?]$", line):  # Sentence-ending punctuation
                            temp_text += line
                            text += temp_text + "\n"
                            temp_text = ""  # Reset
                        else:
                            temp_text += line + " "
                    if temp_text:
                        text += temp_text.strip() + "\n"
    except Exception as e:
        print(f"Error extracting PDF text: {e}")
    return text

# Function to extract text from DOCX
def extract_text_from_docx(file_path):
    text = ""
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            sentences = para.text.split(".")
            for sentence in sentences:
                text += sentence.strip() + "\n"
    except Exception as e:
        print(f"Error extracting DOCX text: {e}")
    return text

def extract_text_from_image(file_path):
    # Set a dynamic path to the Tesseract executable
    base_dir = os.path.dirname(os.path.abspath(__file__))
    tesseract_path = os.path.join(base_dir, "Tesseract-OCR", "tesseract.exe")
    pytesseract.pytesseract.tesseract_cmd = tesseract_path
    try:
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
        if not text:
            raise ValueError("No text could be extracted from the image.")
        cleaned_text = " ".join(text.splitlines())
        cleaned_text = cleaned_text.replace(" .", ".").replace(" ,", ",")
        sentences = re.split(r"([.?!])\s*", cleaned_text)
        formatted_text = ""
        for i in range(0, len(sentences), 2):
            sentence = sentences[i].strip()
            if i + 1 < len(sentences):
                sentence += sentences[i + 1]
            if sentence:
                formatted_text += sentence.strip() + "\n"
        return formatted_text
    except Exception as e:
        print(f"Error in extract_text_from_image: {e}")
        return None

# Function to extract text from txt file
def extract_text_from_txt(file):
    try:
        # Read the file directly from the request
        text = file.read().decode("utf-8")

        # Process text as sentences
        processed_text = process_text(text)
        return processed_text

    except Exception as e:
        print(f"Error extracting TXT text: {e}")
    return ""

# Function to process text into properly formatted sentences
def process_text(text):
    # Split text into sentences using punctuation
    sentences = re.split(r"([.!?])(?=\s|$)", text)
    # Reconstruct sentences with new lines
    formatted_text = "".join(
        [
            sentence.strip() + punctuation + "\n"
            for sentence, punctuation in zip(sentences[::2], sentences[1::2])
        ]
    )
    # Handle trailing sentence without punctuation
    if len(sentences) % 2 == 1 and sentences[-1].strip():
        formatted_text += sentences[-1].strip()
    return formatted_text

# Function to extract text from xlsx file
def extract_text_from_xlsx(file):
    workbook = openpyxl.load_workbook(file)
    extracted_text = ""
    # Loop through each sheet in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            row_text = ""
            # Loop through each cell in the row
            for cell in row:
                if cell == None:
                    row_text += ""
                else:
                    row_text += str(cell)  # Convert cell value to string
                row_text += "\t"  # Add tab between cells
            extracted_text += row_text.strip() + "\n"  # Add row text to extracted text
    return extracted_text

# Function to extract text from CSV file
def extract_text_from_csv(file):
    extracted_text = ""
    try:
        # Ensure the file is in text mode
        file_content = file.stream.read().decode(
            "utf-8"
        )  # Read file content and decode
        file_io = io.StringIO(
            file_content
        )  # Wrap the content in a StringIO object for csv.reader

        reader = csv.reader(file_io)
        for row in reader:
            row_text = ",".join(row)  # Join the row with commas
            extracted_text += row_text + "\n"  # Add newline for each row

    except csv.Error as e:
        return f"Error processing CSV file: {e}"
    except Exception as e:
        return f"An unexpected error occurred: {e}"

    return extracted_text

# Function to extract text from JSON file
def extract_text_from_json(file):
    try:
        # Read the file content as bytes
        file_content = file.read()
        if not file_content:
            print("File is empty")
            return "No content found in the file."
        json_data = json.loads(file_content.decode("utf-8"))
        extracted_text = json.dumps(json_data, indent=4)
        return extracted_text

    except json.JSONDecodeError as e:
        print(f"Error decoding JSON in file: {e}")
        return "Error decoding JSON in the file."
    except Exception as e:
        print(f"Unexpected error: {e}")
        return f"Unexpected error: {e}"
    return all_data

# Function to read a .MD file
def extract_text_from_md(file):
    try:
        # Read the file and decode it into a string
        data = file.read().decode("utf-8")
        # Remove lines that start with a # (markdown headers)
        data = re.sub(r"^\#.*$", "", data, flags=re.M)
        # Remove images (anything starting with ![ and ending with ])
        data = re.sub(r"!\[.*?\]\(.*?\)", "", data)
        # Remove links (anything in the format [text](url))
        data = re.sub(r"\[.*?\]\(.*?\)", "", data)
        # Remove bold or italic text (anything between **, *, __, or _)
        data = re.sub(r"(\*\*|\*|__|_)(.*?)\1", r"\2", data)
        # Remove code blocks (text wrapped in triple backticks)
        data = re.sub(r"```.*?```", "", data, flags=re.S)
        # Remove indented code (with spaces at the beginning)
        data = re.sub(r"    .+", "", data)
        # Remove list items (lines that start with *, +, or -)
        data = re.sub(r"^[\*\+-]\s.*$", "", data, flags=re.M)
        # Remove blockquotes (lines starting with >)
        data = re.sub(r"^\>.*$", "", data, flags=re.M)
        # Strip extra spaces at the beginning and end
        data = data.strip()
        return data

    except Exception as e:
        return "Something went wrong: " + str(e)

# Function to read a .reg file and extract key-value pairs
def read_reg(file):
    try:
        # Read the contents of the .reg file (UTF-16 encoding is common for .reg files)
        data = file.read().decode("utf-16")
        # Remove non-section strings before the first section (if any)
        data = re.sub(r"^[^\[]*\n", "", data, flags=re.S)
        # Find sections and key-value pairs
        sections = re.findall(r"\[([^\]]+)\](.*?)((?=\[)|$)", data, re.S)
        # Prepare a list to store rows of data
        rows = []

        for section, content, _ in sections:
            # Find key-value pairs within the section
            key_value_pairs = re.findall(r"([^\=]+)=(.*)", content.strip())
            for key, value in key_value_pairs:
                rows.append([section.strip(), key.strip(), value.strip()])

        # Convert the list into a DataFrame
        df = pd.DataFrame(rows, columns=["Section", "Key", "Value"])
        # Return the DataFrame
        return df

    except Exception as e:
        print(f"Error reading .reg file: {e}")
        return "Error reading .reg file"

# Route for processing messages
@app.route("/process", methods=["POST"])
def process_message():
    response = ""

    # Case 1: message only 
    if "message" in request.form and not request.files:
        message = request.form["message"]
        # Get GPT response for the message
        gpt_response = get_response(message)
        response += f"\n{gpt_response}"

        # Save the message and response to the database
        save_chat_to_db(message, gpt_response)

    # Case 2: file only 
    elif "file" in request.files and not "message" in request.form:
        file = request.files["file"]
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            extracted_text = ""

            # Extract text from the file
            if filename.endswith(".pdf"):
                extracted_text = extract_text_from_pdf(file)
            elif filename.endswith(".txt"):
                extracted_text = extract_text_from_txt(file)
            elif filename.endswith(".json"):
                extracted_text = extract_text_from_json(file)
            elif filename.endswith(".docx"):
                extracted_text = extract_text_from_docx(file)
            elif filename.endswith(".xlsx"):
                extracted_text = extract_text_from_xlsx(file)
            elif filename.endswith(".csv"):
                extracted_text = extract_text_from_csv(file)
            elif filename.endswith((".jpg", ".jpeg", ".png")):
                extracted_text = extract_text_from_image(file)
            elif filename.endswith(".md"):
                extracted_text = extract_text_from_md(file)
            elif filename.endswith(".reg"):
                df = read_reg(file)
                extracted_text = df.to_html()

            if extracted_text.strip():
                # Get GPT response for the extracted text only
                gpt_response = get_response(extracted_text)
                response += f"\n{gpt_response}"

                # Save the extracted text and GPT response to the database
                save_chat_to_db(extracted_text, gpt_response)
            else:
                response += "No text extracted from the file."
        else:
            return jsonify({"error": "Invalid file type"}), 400

    # Case 3: both a message and a file
    elif "file" in request.files and "message" in request.form:
        file = request.files["file"]
        message = request.form["message"]

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            extracted_text = ""

            # Extract text from the file
            if filename.endswith(".pdf"):
                extracted_text = extract_text_from_pdf(file)
            elif filename.endswith(".txt"):
                extracted_text = extract_text_from_txt(file)
            elif filename.endswith(".json"):
                extracted_text = extract_text_from_json(file)
            elif filename.endswith(".docx"):
                extracted_text = extract_text_from_docx(file)
            elif filename.endswith(".xlsx"):
                extracted_text = extract_text_from_xlsx(file)
            elif filename.endswith(".csv"):
                extracted_text = extract_text_from_csv(file)
            elif filename.endswith((".jpg", ".jpeg", ".png")):
                extracted_text = extract_text_from_image(file)
            elif filename.endswith(".md"):
                extracted_text = extract_text_from_md(file)
            elif filename.endswith(".reg"):
                df = read_reg(file)
                extracted_text = df.to_html()

            if extracted_text.strip():
                # Combine the message and extracted text
                final_input = message + "\n" + extracted_text

                # Get GPT response based on the combined input
                gpt_response = get_response(final_input)
                response += f"{gpt_response}"

                # Save the final input and GPT response to the database
                save_chat_to_db(final_input, gpt_response)
            else:
                response += "No text extracted from the file."
        else:
            return jsonify({"error": "Invalid file type"}), 400

    return jsonify({"response": response})

# Save chat to the database
def save_chat_to_db(message, gpt_response):
    user_id = session.get("user_id")
    session_id = session.get("session_id")

    if user_id and session_id:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Insert the message and response into the messages table
            cursor.execute(
                "INSERT INTO messages (session_id, message, response) VALUES (?, ?, ?)",
                (session_id, message, gpt_response),
            )
            conn.commit()
            conn.close()

        except Exception as e:
            print(f"Error saving chat to database: {e}")
            conn.close()

# Route to generate a report
@app.route("/report", methods=["POST"])
def process_report():
    report = ""

    if "file" in request.files and not "message" in request.form:
        file = request.files["file"]

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            extracted_text = ""

            # Extract text from the file
            if filename.endswith(".pdf"):
                extracted_text = extract_text_from_pdf(file)
            elif filename.endswith(".txt"):
                extracted_text = extract_text_from_txt(file)
            elif filename.endswith(".json"):
                extracted_text = extract_text_from_json(file)
            elif filename.endswith(".docx"):
                extracted_text = extract_text_from_docx(file)
            elif filename.endswith(".xlsx"):
                extracted_text = extract_text_from_xlsx(file)
            elif filename.endswith(".csv"):
                extracted_text = extract_text_from_csv(file)
            elif filename.endswith((".jpg", ".jpeg", ".png")):
                extracted_text = extract_text_from_image(file)
            elif filename.endswith(".md"):
                extracted_text = extract_text_from_md(file)
            elif filename.endswith(".reg"):
                df = read_reg(file)
                extracted_text = df.to_html()

            if extracted_text.strip():
                # Get GPT response for the extracted text only
                gpt_response = get_report(extracted_text)
                report += f"\n{gpt_response}"
            else:
                report += "No text extracted from the file."
        else:
            return jsonify({"error": "Invalid file type"}), 400

        # Create a BytesIO object for the report content
        report_file = BytesIO()
        report_file.write(report.encode('utf-8'))  # Encode string to bytes
        report_file.seek(0)

        # Send the text file to the user
        return send_file(
            report_file,
            as_attachment=True,
            download_name="report.txt",
            mimetype="text/plain"
        )

# Route to get chat history
@app.route("/get_chat_history", methods=["GET"])
def get_chat_history():
    user_id = session.get("user_id")
    session_name_filter = request.args.get("session_name")  # New query parameter to filter by session name

    if not user_id:
        return jsonify({"success": False, "message": "User not logged in"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # If session_name_filter is provided, filter by session name
        if session_name_filter:
            cursor.execute(
                """
                SELECT cs.session_id, cs.session_name, m.message, m.response, m.timestamp
                FROM chat_sessions cs
                LEFT JOIN messages m ON cs.session_id = m.session_id
                WHERE cs.user_id = ? AND cs.session_name = ?
                ORDER BY cs.session_name DESC, m.timestamp DESC
                """, (user_id, session_name_filter)
            )
        else:
            cursor.execute(
                """
                SELECT cs.session_id, cs.session_name, m.message, m.response, m.timestamp
                FROM chat_sessions cs
                LEFT JOIN messages m ON cs.session_id = m.session_id
                WHERE cs.user_id = ?
                ORDER BY cs.session_name DESC, m.timestamp DESC
                """, (user_id,)
            )

        rows = cursor.fetchall()

        chat_history = []
        for row in rows:
            chat_history.append({
                "session_id": row[0],
                "session_name": row[1],
                "message": row[2],
                "response": row[3],
                "timestamp": row[4]
            })

        cursor.close()
        conn.close()

        return jsonify({
            "success": True,
            "chat_history": chat_history
        })

    except Exception as e:
        print(f"Error getting chat history: {e}")
        
        if conn:
            cursor.close()
            conn.close()

        return jsonify({"success": False, "message": "Error getting chat history"})

# route to clear chat history 
@app.route("/clear_chat_history", methods=["DELETE"])
def clear_chat_history():
    user_id = session.get("user_id")
    current_session_id = session.get("session_id")  
    session_name_filter = request.args.get("session_name")

    if not user_id:
        return jsonify({"success": False, "message": "User not logged in"})

    if not session_name_filter:
        return jsonify({"success": False, "message": "Session name must be provided"})

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Step 1: Get the session_id for the provided session_name
        cursor.execute(
            """
            SELECT session_id
            FROM chat_sessions
            WHERE user_id = ? AND session_name = ?
            """, (user_id, session_name_filter)
        )

        session_record = cursor.fetchone()
        
        # If no session is found, return an error
        if not session_record:
            return jsonify({"success": False, "message": "Session not found"})

        session_id = session_record[0]

        # Check if the session being deleted is the current session
        if session_id == current_session_id:
            return jsonify({
                "success": False,
                "message": "Cannot delete the current active session"
            })

        # Step 2: Delete messages associated with the session_name
        cursor.execute(
            """
            DELETE FROM messages
            WHERE session_id = ?
            """, (session_id,)
        )

        # Step 3: Delete the session record itself
        cursor.execute(
            """
            DELETE FROM chat_sessions
            WHERE session_id = ?
            """, (session_id,)
        )

        # Commit changes to the database
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": f"Session '{session_name_filter}' has been deleted"
        })

    except Exception as e:
        print(f"Error deleting session: {e}")

        if conn:
            cursor.close()
            conn.close()

        return jsonify({"success": False, "message": f"Error deleting session: {str(e)}"})

# Route to export chat history
@app.route("/export_chat_history", methods=["GET"])
def export_chat_history():
    user_id = session.get("user_id")
    session_name_filter = request.args.get("session_name")

    if not user_id:
        return jsonify({"error": "User not logged in or no active session"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if session_name_filter:
            cursor.execute(
                "SELECT * FROM chat_sessions WHERE user_id = ? AND session_name = ?", 
                (user_id, session_name_filter)
            )
        else:
            cursor.execute("SELECT * FROM chat_sessions WHERE user_id = ?", (user_id,))

        db_session = cursor.fetchone()

        if not db_session:
            return jsonify({"error": "Session not found for the current user"}), 404

        session_id = db_session['session_id']

        cursor.execute(
            "SELECT id, message, response, datetime(timestamp, '+8 hours') AS timestamp_utc8 FROM messages WHERE session_id = ? ORDER BY timestamp", 
            (session_id,)
        )
        messages = cursor.fetchall()

        export_data = {
            "session_id": session_id,
            "session_name": db_session["session_name"],
            "user_id": user_id,
            "timestamp": db_session["timestamp"],
            "messages": [
                {
                    "id": message["id"],
                    "message": message["message"],
                    "response": message["response"],
                    "timestamp": message["timestamp_utc8"]
                }
                for message in messages
            ]
        }

        json_data = json.dumps(export_data, default=str, indent=4)

        byte_io = BytesIO(json_data.encode('utf-8'))
        byte_io.seek(0)

        return send_file(
            byte_io,
            as_attachment=True,
            download_name="chat_history.json",
            mimetype='application/json'
        )

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": f"Error generating the chat history file: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()

# Route to logout
@app.route("/logout")
def logout():
    session.clear()
    return jsonify({"message": "You have successfully logged out."}), 200

# Function to delete account
@app.route("/delete_account", methods=["DELETE"])
def delete_account():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "User not logged in"}), 400

    try:
        # Connect to the database
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if the user exists
        cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
        user = cursor.fetchone()

        if not user:
            conn.close()
            return jsonify({"error": "User not found"}), 404

        # Delete all associated chat messages
        cursor.execute("DELETE FROM messages WHERE session_id IN (SELECT session_id FROM chat_sessions WHERE user_id = ?)", (user_id,))
        # Delete all associated chat sessions
        cursor.execute("DELETE FROM chat_sessions WHERE user_id = ?", (user_id,))
        # Delete the user from the users table
        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))

        # Commit the changes
        conn.commit()
        conn.close()

        # Clear session
        session.clear()

        return jsonify({"success": True, "message": "Account and all associated data deleted successfully"}), 200

    except Exception as e:
        print(f"Error during account deletion: {e}")
        return jsonify({"error": "An error occurred during deletion. Please try again later."}), 500

# Main entry point
if __name__ == "__main__":
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    init_db()
    app.run(debug=True, host="127.0.0.1", port=5000)